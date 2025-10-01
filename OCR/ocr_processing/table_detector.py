"""
Advanced Table Detection and Structure Extraction Module
Implements robust table detection using OpenCV line detection and grid analysis
"""
import cv2
import numpy as np
import pytesseract
from PIL import Image
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass
import logging

logger = logging.getLogger(__name__)


@dataclass
class CellInfo:
    """Information about a detected table cell"""
    row: int
    col: int
    x: int
    y: int
    width: int
    height: int
    text: str = ""
    confidence: float = 0.0
    is_header: bool = False


@dataclass
class TableStructure:
    """Complete table structure with cells and metadata"""
    rows: int
    cols: int
    cells: List[CellInfo]
    headers: Dict[str, str]  # {col_index: header_text}
    grid_confidence: float
    

class TableDetector:
    """Detects and extracts table structure from images"""
    
    def __init__(self, ocr_engine=None):
        """
        Initialize table detector
        
        Args:
            ocr_engine: Optional OCR engine instance (defaults to pytesseract)
        """
        self.ocr_engine = ocr_engine
        
    def preprocess_image(self, image: np.ndarray) -> np.ndarray:
        """
        Preprocess image for better table detection
        
        Args:
            image: Input image (BGR or grayscale)
            
        Returns:
            Preprocessed image ready for line detection
        """
        # Convert to grayscale if needed
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy()
        
        # Apply adaptive thresholding
        binary = cv2.adaptiveThreshold(
            gray, 255, 
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY_INV, 
            11, 2
        )
        
        return binary
    
    def detect_lines(self, binary_image: np.ndarray) -> Tuple[List, List]:
        """
        Detect horizontal and vertical lines in the image
        
        Args:
            binary_image: Binary image with table
            
        Returns:
            Tuple of (horizontal_lines, vertical_lines)
        """
        # Detect horizontal lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        horizontal_lines = cv2.morphologyEx(
            binary_image, 
            cv2.MORPH_OPEN, 
            horizontal_kernel, 
            iterations=2
        )
        
        # Detect vertical lines
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
        vertical_lines = cv2.morphologyEx(
            binary_image, 
            cv2.MORPH_OPEN, 
            vertical_kernel, 
            iterations=2
        )
        
        # Find contours to get line positions
        h_contours, _ = cv2.findContours(
            horizontal_lines, 
            cv2.RETR_EXTERNAL, 
            cv2.CHAIN_APPROX_SIMPLE
        )
        
        v_contours, _ = cv2.findContours(
            vertical_lines, 
            cv2.RETR_EXTERNAL, 
            cv2.CHAIN_APPROX_SIMPLE
        )
        
        # Extract y-coordinates of horizontal lines
        h_lines = []
        for contour in h_contours:
            x, y, w, h = cv2.boundingRect(contour)
            if w > 50:  # Filter out noise
                h_lines.append(y)
        
        # Extract x-coordinates of vertical lines
        v_lines = []
        for contour in v_contours:
            x, y, w, h = cv2.boundingRect(contour)
            if h > 50:  # Filter out noise
                v_lines.append(x)
        
        # Sort lines
        h_lines = sorted(set(h_lines))
        v_lines = sorted(set(v_lines))
        
        logger.info(f"Detected {len(h_lines)} horizontal lines and {len(v_lines)} vertical lines")
        
        return h_lines, v_lines
    
    def detect_grid_with_hough(self, binary_image: np.ndarray) -> Tuple[List, List]:
        """
        Alternative method: Use Hough Line Transform for line detection
        
        Args:
            binary_image: Binary image with table
            
        Returns:
            Tuple of (horizontal_lines, vertical_lines)
        """
        # Apply edge detection
        edges = cv2.Canny(binary_image, 50, 150, apertureSize=3)
        
        # Detect lines using Hough Transform
        lines = cv2.HoughLinesP(
            edges,
            rho=1,
            theta=np.pi/180,
            threshold=100,
            minLineLength=100,
            maxLineGap=10
        )
        
        if lines is None:
            logger.warning("No lines detected with Hough Transform")
            return [], []
        
        h_lines = []
        v_lines = []
        
        for line in lines:
            x1, y1, x2, y2 = line[0]
            
            # Check if line is horizontal (small y difference)
            if abs(y2 - y1) < 10:
                h_lines.append((y1 + y2) // 2)
            
            # Check if line is vertical (small x difference)
            elif abs(x2 - x1) < 10:
                v_lines.append((x1 + x2) // 2)
        
        # Remove duplicates and sort
        h_lines = sorted(set(h_lines))
        v_lines = sorted(set(v_lines))
        
        logger.info(f"Hough detected {len(h_lines)} horizontal, {len(v_lines)} vertical lines")
        
        return h_lines, v_lines
    
    def build_grid(self, h_lines: List[int], v_lines: List[int]) -> List[CellInfo]:
        """
        Build cell grid from detected lines
        
        Args:
            h_lines: List of y-coordinates for horizontal lines
            v_lines: List of x-coordinates for vertical lines
            
        Returns:
            List of CellInfo objects representing the grid
        """
        cells = []
        
        # Create cells from intersections
        for row_idx in range(len(h_lines) - 1):
            for col_idx in range(len(v_lines) - 1):
                y1 = h_lines[row_idx]
                y2 = h_lines[row_idx + 1]
                x1 = v_lines[col_idx]
                x2 = v_lines[col_idx + 1]
                
                # Create cell
                cell = CellInfo(
                    row=row_idx,
                    col=col_idx,
                    x=x1,
                    y=y1,
                    width=x2 - x1,
                    height=y2 - y1,
                    is_header=(row_idx == 0)  # First row is header
                )
                
                cells.append(cell)
        
        logger.info(f"Built grid with {len(cells)} cells")
        
        return cells
    
    def extract_cell_text(self, image: np.ndarray, cell: CellInfo) -> Tuple[str, float]:
        """
        Extract text from a specific cell using OCR
        
        Args:
            image: Original image
            cell: Cell information with coordinates
            
        Returns:
            Tuple of (extracted_text, confidence)
        """
        # Add padding to avoid cutting text
        padding = 5
        y1 = max(0, cell.y + padding)
        y2 = min(image.shape[0], cell.y + cell.height - padding)
        x1 = max(0, cell.x + padding)
        x2 = min(image.shape[1], cell.x + cell.width - padding)
        
        # Extract cell region
        cell_img = image[y1:y2, x1:x2]
        
        if cell_img.size == 0:
            return "", 0.0
        
        try:
            # Run OCR on cell
            if self.ocr_engine:
                # Use provided OCR engine
                from ocr_processing.ocr_core import OCRResult
                result = self.ocr_engine.extract_text_tesseract(cell_img)
                return result.text.strip(), result.confidence
            else:
                # Use pytesseract directly
                data = pytesseract.image_to_data(
                    cell_img, 
                    output_type=pytesseract.Output.DICT
                )
                
                # Extract text and confidence
                text = " ".join([
                    word for word, conf in zip(data['text'], data['conf'])
                    if conf > 0 and word.strip()
                ])
                
                # Calculate average confidence
                confidences = [c for c in data['conf'] if c > 0]
                avg_conf = sum(confidences) / len(confidences) if confidences else 0.0
                
                return text.strip(), avg_conf
                
        except Exception as e:
            logger.error(f"Error extracting text from cell: {e}")
            return "", 0.0
    
    def detect_table_structure(
        self, 
        image_path: str, 
        method: str = "morphology"
    ) -> Optional[TableStructure]:
        """
        Main method: Detect complete table structure from image
        
        Args:
            image_path: Path to image file
            method: Detection method ("morphology" or "hough")
            
        Returns:
            TableStructure object or None if detection fails
        """
        try:
            # Load image
            image = cv2.imread(image_path)
            if image is None:
                logger.error(f"Failed to load image: {image_path}")
                return None
            
            # Preprocess
            binary = self.preprocess_image(image)
            
            # Detect lines
            if method == "hough":
                h_lines, v_lines = self.detect_grid_with_hough(binary)
            else:
                h_lines, v_lines = self.detect_lines(binary)
            
            # Need at least 2 lines in each direction for a table
            if len(h_lines) < 2 or len(v_lines) < 2:
                logger.warning(f"Insufficient lines detected: {len(h_lines)}h, {len(v_lines)}v")
                return None
            
            # Build grid
            cells = self.build_grid(h_lines, v_lines)
            
            # Extract text from each cell
            for cell in cells:
                text, confidence = self.extract_cell_text(image, cell)
                cell.text = text
                cell.confidence = confidence
            
            # Extract headers (first row)
            headers = {}
            header_cells = [c for c in cells if c.row == 0]
            for cell in header_cells:
                if cell.text:
                    headers[str(cell.col)] = cell.text
            
            # Calculate grid confidence
            all_confidences = [c.confidence for c in cells if c.confidence > 0]
            grid_confidence = (
                sum(all_confidences) / len(all_confidences) 
                if all_confidences else 0.0
            )
            
            # Create structure
            structure = TableStructure(
                rows=len(h_lines) - 1,
                cols=len(v_lines) - 1,
                cells=cells,
                headers=headers,
                grid_confidence=grid_confidence
            )
            
            logger.info(
                f"Detected table: {structure.rows}x{structure.cols}, "
                f"confidence: {grid_confidence:.1f}%"
            )
            
            return structure
            
        except Exception as e:
            logger.error(f"Error detecting table structure: {e}", exc_info=True)
            return None
    
    def structure_to_dict(self, structure: TableStructure) -> Dict[str, Any]:
        """
        Convert TableStructure to dictionary for JSON serialization
        
        Args:
            structure: TableStructure object
            
        Returns:
            Dictionary representation
        """
        return {
            'rows': structure.rows,
            'cols': structure.cols,
            'headers': structure.headers,
            'grid_confidence': structure.grid_confidence,
            'cells': [
                {
                    'row': cell.row,
                    'col': cell.col,
                    'x': cell.x,
                    'y': cell.y,
                    'width': cell.width,
                    'height': cell.height,
                    'text': cell.text,
                    'confidence': cell.confidence,
                    'is_header': cell.is_header
                }
                for cell in structure.cells
            ],
            'field_names': list(structure.headers.values())
        }
    
    def export_to_excel_template(
        self, 
        structure: TableStructure, 
        output_path: str
    ) -> bool:
        """
        Export detected table structure to Excel template
        
        Args:
            structure: TableStructure object
            output_path: Path to save Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            
            # Fill headers (first row)
            for col_idx, header_text in structure.headers.items():
                col = int(col_idx) + 1  # Excel is 1-indexed
                cell = ws.cell(row=1, column=col)
                cell.value = header_text
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", 
                    end_color="CCCCCC", 
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            for col_idx in range(structure.cols):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = 20
            
            # Add empty rows for data entry
            for row_idx in range(2, structure.rows + 1):
                for col_idx in range(1, structure.cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = ""
                    cell.alignment = Alignment(vertical="center")
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Exported Excel template to: {output_path}")
            
            return True
            
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            return False


def visualize_table_detection(
    image_path: str, 
    structure: TableStructure, 
    output_path: str
) -> bool:
    """
    Visualize detected table structure by drawing grid on image
    
    Args:
        image_path: Path to original image
        structure: Detected table structure
        output_path: Path to save visualization
        
    Returns:
        True if successful
    """
    try:
        import cv2
        
        # Load image
        image = cv2.imread(image_path)
        if image is None:
            return False
        
        # Draw grid
        for cell in structure.cells:
            # Draw rectangle
            color = (0, 255, 0) if cell.is_header else (255, 0, 0)
            cv2.rectangle(
                image,
                (cell.x, cell.y),
                (cell.x + cell.width, cell.y + cell.height),
                color,
                2
            )
            
            # Draw text (if detected)
            if cell.text:
                cv2.putText(
                    image,
                    cell.text[:20],  # Truncate long text
                    (cell.x + 5, cell.y + 20),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    (0, 0, 255),
                    1
                )
        
        # Save visualization
        cv2.imwrite(output_path, image)
        logger.info(f"Saved visualization to: {output_path}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error creating visualization: {e}")
        return False
