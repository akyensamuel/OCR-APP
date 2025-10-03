"""
Excel Template Operations for Document Processing
Handles filling Excel templates with extracted OCR data and managing multi-document exports
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from django.conf import settings


class ExcelTemplateManager:
    """Manage Excel template operations for document processing"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def fill_template_with_document_data(self, template_excel_path, document_data, output_path):
        """
        Fill an Excel template with data from a processed document
        
        Args:
            template_excel_path: Path to the Excel template file
            document_data: Dictionary with extracted data from document
            output_path: Path where the filled Excel file will be saved
        
        Returns:
            Path to the saved Excel file
        """
        try:
            # Load the template
            wb = load_workbook(template_excel_path)
            ws = wb.active
            
            # Find the first empty row (skip header row)
            next_row = 2
            while ws.cell(row=next_row, column=1).value is not None:
                next_row += 1
            
            # Fill the row with document data
            if 'fields' in document_data:
                # Template-based extraction with fields
                for i, field in enumerate(document_data['fields']):
                    col_index = i + 1
                    value = field.get('value', '')
                    ws.cell(row=next_row, column=col_index, value=value)
                    
                    # Apply basic styling
                    cell = ws.cell(row=next_row, column=col_index)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = self.thin_border
            
            elif 'cells' in document_data:
                # Table detection extraction
                cells = document_data['cells']
                for cell_data in cells:
                    row = cell_data.get('row', 0) + 1  # +1 for header
                    col = cell_data.get('col', 0) + 1  # Excel is 1-indexed
                    text = cell_data.get('text', '')
                    
                    # Skip header row (row 0)
                    if row > 1:
                        ws.cell(row=row + 1, column=col, value=text)
                        cell = ws.cell(row=row + 1, column=col)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = self.thin_border
            
            # Save the workbook
            wb.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"Error filling Excel template: {str(e)}")
    
    def append_document_to_excel(self, excel_path, document_data):
        """
        Append a new document's data as a new row in an existing Excel file
        
        Args:
            excel_path: Path to the existing Excel file
            document_data: Dictionary with extracted data from document
        
        Returns:
            Path to the updated Excel file
        """
        try:
            # Check if file exists
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
            # Load the workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            # Extract data based on format
            if 'fields' in document_data:
                # Template-based format
                for i, field in enumerate(document_data['fields']):
                    col_index = i + 1
                    value = field.get('value', '')
                    cell = ws.cell(row=next_row, column=col_index, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = self.thin_border
            
            elif 'cells' in document_data:
                # Table detection format - append ALL rows (skip header row 0)
                cells = document_data['cells']
                # Group cells by row
                row_data = {}
                for cell_data in cells:
                    row = cell_data.get('row', 0)
                    col = cell_data.get('col', 0)
                    text = cell_data.get('text', '')
                    if row not in row_data:
                        row_data[row] = {}
                    row_data[row][col] = text
                
                # Append all data rows (skip header at row 0)
                current_row = next_row
                for row_idx in sorted(row_data.keys()):
                    if row_idx > 0:  # Skip header row (row 0)
                        for col_idx in sorted(row_data[row_idx].keys()):
                            cell = ws.cell(row=current_row, column=col_idx + 1, value=row_data[row_idx][col_idx])
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = self.thin_border
                        current_row += 1
            
            # Save the workbook
            wb.save(excel_path)
            return excel_path
            
        except Exception as e:
            raise Exception(f"Error appending to Excel file: {str(e)}")
    
    def create_multi_document_export(self, template, documents, output_path, user_filename=None):
        """
        Create a consolidated Excel file with data from multiple documents
        
        Args:
            template: Template model instance
            documents: List of Document model instances
            output_path: Path where the Excel file will be saved
            user_filename: Optional custom filename from user
        
        Returns:
            Path to the created Excel file
        """
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Set worksheet title
            ws.title = template.name[:31] if len(template.name) <= 31 else template.name[:28] + "..."
            
            # Get headers from template
            headers = template.get_field_names()
            
            if not headers:
                raise ValueError("Template has no field names defined")
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.thin_border
            
            # Write document data
            current_row = 2
            for document in documents:
                extracted_data = document.extracted_data
                
                if 'fields' in extracted_data:
                    # Template-based format
                    for col_idx, field in enumerate(extracted_data['fields'], start=1):
                        value = field.get('value', '')
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = self.thin_border
                
                elif 'cells' in extracted_data:
                    # Table detection format - extract ALL data rows (skip header row 0)
                    cells = extracted_data['cells']
                    row_data = {}
                    for cell_data in cells:
                        row = cell_data.get('row', 0)
                        col = cell_data.get('col', 0)
                        text = cell_data.get('text', '')
                        if row not in row_data:
                            row_data[row] = {}
                        row_data[row][col] = text
                    
                    # Export all data rows (skip header at row 0)
                    for row_idx in sorted(row_data.keys()):
                        if row_idx > 0:  # Skip header row
                            for col_idx in sorted(row_data[row_idx].keys()):
                                if col_idx + 1 <= len(headers):
                                    cell = ws.cell(row=current_row, column=col_idx + 1, value=row_data[row_idx][col_idx])
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                    cell.border = self.thin_border
                            current_row += 1
                
                else:
                    # No data, just increment row
                    current_row += 1
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Determine filename
            if user_filename:
                if not user_filename.endswith('.xlsx'):
                    user_filename += '.xlsx'
                final_path = os.path.join(os.path.dirname(output_path), user_filename)
            else:
                final_path = output_path
            
            # Save the workbook
            wb.save(final_path)
            return final_path
            
        except Exception as e:
            raise Exception(f"Error creating multi-document export: {str(e)}")
    
    def get_template_excel_path(self, template):
        """
        Get the path to the Excel template file for a given template
        
        Args:
            template: Template model instance
        
        Returns:
            Path to the Excel template file, or None if not found
        """
        if not template.file:
            return None
        
        # Get the base path and construct Excel template path
        file_path = template.file.path
        base_path, ext = os.path.splitext(file_path)
        excel_path = f"{base_path}_template.xlsx"
        
        if os.path.exists(excel_path):
            return excel_path
        
        return None
