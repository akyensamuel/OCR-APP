from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import Document
from templates.models import Template


def document_list(request):
    """List all processed documents"""
    documents = Document.objects.all().order_by('-created_at')
    return render(request, 'documents/document_list.html', {'documents': documents})


def document_upload(request):
    """Upload a document for general OCR processing"""
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('document_file')
            if not uploaded_file:
                messages.error(request, 'No file selected')
                return redirect('documents:document_upload')
            
            # Process with OCR
            from ocr_processing.ocr_core import OCREngine
            from basemode.file_storage import save_file_to_db, get_temp_file_path, cleanup_temp_file
            import os
            
            # Save file to database
            file_info = save_file_to_db(uploaded_file)
            
            # Create temp file for OCR processing
            full_path = get_temp_file_path(file_info['file_data'], file_info['file_name'])
            
            # Initialize OCR and extract text
            ocr_engine = OCREngine()
            ocr_result = ocr_engine.extract_text(full_path)
            
            # Get current user or create default user
            if request.user.is_authenticated:
                uploaded_by = request.user
            else:
                from django.contrib.auth.models import User
                uploaded_by, created = User.objects.get_or_create(
                    username='anonymous',
                    defaults={
                        'email': 'anonymous@example.com',
                        'first_name': 'Anonymous',
                        'last_name': 'User'
                    }
                )
            
            # Create Document record with database storage
            document = Document.objects.create(
                name=file_info['file_name'],
                file_data=file_info['file_data'],
                file_name=file_info['file_name'],
                file_type=file_info['file_type'],
                file_size=file_info['file_size'],
                uploaded_by=uploaded_by,
                extracted_data={
                    'text': ocr_result.text,
                    'confidence': ocr_result.confidence,
                    'engine': ocr_result.engine
                },
                processing_status='completed'
            )
            
            # Cleanup temp file
            cleanup_temp_file(full_path)
            
            messages.success(request, f'Document processed successfully! Confidence: {ocr_result.confidence:.1f}%')
            return redirect('documents:document_detail', document_id=document.pk)
            
        except Exception as e:
            # Cleanup temp file on error
            try:
                if 'full_path' in locals():
                    cleanup_temp_file(full_path)
            except:
                pass
            messages.error(request, f'Error processing document: {str(e)}')
            return redirect('documents:document_upload')
    
    return render(request, 'documents/document_upload.html')


def document_upload_with_template(request, template_id):
    """Upload a document using a specific template"""
    template = get_object_or_404(Template, id=template_id)
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('document_file')
            if not uploaded_file:
                messages.error(request, 'No file selected')
                return redirect('documents:document_upload_with_template', template_id=template_id)
            
            # Process with template
            from ocr_processing.ocr_core import OCREngine
            from ocr_processing.table_detector import TableDetector
            from ocr_processing.excel_manager import ExcelTemplateManager
            from basemode.file_storage import save_file_to_db, get_temp_file_path, cleanup_temp_file, read_file_from_path
            import os
            
            # Save uploaded file to database
            file_info = save_file_to_db(uploaded_file)
            
            # Create temp file for OCR processing
            full_path = get_temp_file_path(file_info['file_data'], file_info['file_name'])
            
            # Get current user or create default user
            if request.user.is_authenticated:
                uploaded_by = request.user
            else:
                from django.contrib.auth.models import User
                uploaded_by, created = User.objects.get_or_create(
                    username='anonymous',
                    defaults={
                        'email': 'anonymous@example.com',
                        'first_name': 'Anonymous',
                        'last_name': 'User'
                    }
                )
            
            # Initialize OCR
            ocr_engine = OCREngine()
            
            # Check if template has table structure (new detection method)
            has_table_structure = (
                template.structure and 
                ('headers' in template.structure or 'cells' in template.structure)
            )
            
            extracted_data = {}
            excel_file_path = None
            
            if has_table_structure:
                # Use table detection for processing
                table_detector = TableDetector(ocr_engine)
                table_structure = table_detector.detect_table_structure(full_path, method="morphology")
                
                if table_structure:
                    extracted_data = table_detector.structure_to_dict(table_structure)
                    
                    # Try to fill Excel template if it exists
                    excel_manager = ExcelTemplateManager()
                    template_excel_path = excel_manager.get_template_excel_path(template)
                    
                    if template_excel_path:
                        # Generate Excel in temp location
                        base_name = os.path.splitext(file_info['file_name'])[0]
                        output_excel_name = f"{base_name}_extracted.xlsx"
                        import tempfile
                        excel_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                        output_excel_path = excel_temp.name
                        excel_temp.close()
                        
                        # Append data to Excel template
                        excel_manager.append_document_to_excel(template_excel_path, extracted_data)
                        
                        # Copy template to temp file
                        import shutil
                        shutil.copy(template_excel_path, output_excel_path)
                        
                        # Read Excel file to binary for database storage
                        excel_info = read_file_from_path(output_excel_path)
                        excel_file_data = excel_info['file_data']
                        excel_file_name = output_excel_name
                        
                        # Cleanup temp Excel file
                        cleanup_temp_file(output_excel_path)
                else:
                    # No table detected in document - use fallback extraction
                    messages.warning(request, 'No table structure detected in document. Using fallback extraction.')
                    from ocr_processing.ocr_core import TemplateProcessor
                    template_processor = TemplateProcessor(ocr_engine)
                    extracted_fields = template_processor.process_document_with_template(
                        full_path, template.structure or {}
                    )
                    extracted_data = {
                        'fields': [
                            {
                                'name': field.name,
                                'value': field.value,
                                'confidence': field.confidence
                            } for field in extracted_fields
                        ]
                    }
            
            else:
                # Fallback to old template processor
                from ocr_processing.ocr_core import TemplateProcessor
                template_processor = TemplateProcessor(ocr_engine)
                extracted_fields = template_processor.process_document_with_template(
                    full_path, template.structure or {}
                )
                extracted_data = {
                    'fields': [
                        {
                            'name': field.name,
                            'value': field.value,
                            'confidence': field.confidence
                        } for field in extracted_fields
                    ]
                }
            
            # Create Document record with database storage
            document = Document.objects.create(
                name=file_info['file_name'],
                file_data=file_info['file_data'],
                file_name=file_info['file_name'],
                file_type=file_info['file_type'],
                file_size=file_info['file_size'],
                template=template,
                uploaded_by=uploaded_by,
                extracted_data=extracted_data,
                processing_status='completed'
            )
            
            # Attach Excel file data if created
            if 'excel_file_data' in locals():
                document.excel_data = excel_file_data
                document.excel_name = excel_file_name
                document.save()
            
            # Cleanup temp source file
            cleanup_temp_file(full_path)
            
            field_count = len(extracted_data.get('fields', [])) or len(extracted_data.get('cells', []))
            messages.success(request, f'Document processed with template "{template.name}". Extracted {field_count} items.')
            return redirect('documents:document_detail', document_id=document.pk)
            
        except Exception as e:
            # Cleanup temp files on error
            try:
                if 'full_path' in locals():
                    cleanup_temp_file(full_path)
                if 'output_excel_path' in locals():
                    cleanup_temp_file(output_excel_path)
            except:
                pass
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error processing document: {str(e)}')
            return redirect('documents:document_upload_with_template', template_id=template_id)
    
    return render(request, 'documents/document_upload_template.html', {'template': template})


def document_detail(request, document_id):
    """View document details and extracted data"""
    document = get_object_or_404(Document, id=document_id)
    return render(request, 'documents/document_detail.html', {'document': document})


def document_edit(request, document_id):
    """Edit extracted document data"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        try:
            # Handle different types of extracted data
            if document.extracted_data and 'cells' in document.extracted_data:
                # Table detection format: edit individual cells
                updated_cells = []
                cells = document.extracted_data['cells']
                
                for i, cell in enumerate(cells):
                    # Get the new value from POST data
                    new_value = request.POST.get(f'cell_{i}_value', cell.get('text', ''))
                    
                    updated_cells.append({
                        'row': cell.get('row', 0),
                        'col': cell.get('col', 0),
                        'text': new_value,
                        'confidence': cell.get('confidence', 0),
                        'bbox': cell.get('bbox', []),
                        'manually_edited': True
                    })
                
                document.extracted_data['cells'] = updated_cells
                document.extracted_data['last_edited'] = str(timezone.now())
                
                messages.success(request, f'Document cells updated successfully. Modified {len(updated_cells)} cells.')
                
            elif document.template and document.extracted_data and 'fields' in document.extracted_data:
                # Template-based document: edit individual fields
                updated_fields = []
                
                for i, field in enumerate(document.extracted_data['fields']):
                    field_name = field.get('name', f'field_{i}')
                    new_value = request.POST.get(f'field_{i}_value', field.get('value', ''))
                    
                    updated_fields.append({
                        'name': field_name,
                        'value': new_value,
                        'confidence': field.get('confidence', 0),
                        'manually_edited': True
                    })
                
                document.extracted_data['fields'] = updated_fields
                document.extracted_data['last_edited'] = str(timezone.now())
                
                messages.success(request, f'Document fields updated successfully. Modified {len(updated_fields)} fields.')
                
            else:
                # General text document: edit raw text
                new_text = request.POST.get('extracted_text', '')
                if document.extracted_data:
                    document.extracted_data['text'] = new_text
                    document.extracted_data['manually_edited'] = True
                    document.extracted_data['last_edited'] = str(timezone.now())
                else:
                    document.extracted_data = {
                        'text': new_text,
                        'manually_edited': True,
                        'last_edited': str(timezone.now())
                    }
                
                messages.success(request, 'Document text updated successfully.')
            
            document.save()
            return redirect('documents:document_detail', document_id=document_id)
            
        except Exception as e:
            messages.error(request, f'Error updating document: {str(e)}')
    
    return render(request, 'documents/document_edit.html', {'document': document})


def document_delete(request, document_id):
    """Delete a document"""
    document = get_object_or_404(Document, id=document_id)
    if request.method == 'POST':
        document.delete()
        messages.success(request, f'Document "{document.name}" has been deleted.')
        return redirect('documents:document_list')
    return render(request, 'documents/document_delete.html', {'document': document})


def document_reprocess(request, document_id):
    """Reprocess a document with OCR"""
    document = get_object_or_404(Document, id=document_id)
    if request.method == 'POST':
        try:
            from ocr_processing.ocr_core import OCREngine, TemplateProcessor
            from ocr_processing.table_detector import TableDetector
            from django.conf import settings
            import os
            
            # Get full file path
            full_path = os.path.join(settings.MEDIA_ROOT, document.file.name)
            
            # Initialize OCR engine
            ocr_engine = OCREngine()
            
            if document.template:
                # Check if template has table structure (new detection method)
                has_table_structure = (
                    document.template.structure and 
                    ('headers' in document.template.structure or 'cells' in document.template.structure)
                )
                
                if has_table_structure:
                    # Use table detection for processing
                    table_detector = TableDetector(ocr_engine)
                    table_structure = table_detector.detect_table_structure(full_path, method="morphology")
                    
                    if table_structure:
                        document.extracted_data = table_detector.structure_to_dict(table_structure)
                        cell_count = len(document.extracted_data.get('cells', []))
                        messages.success(request, f'Document reprocessed with table detection. Extracted {cell_count} cells.')
                    else:
                        # No table detected - use fallback
                        messages.warning(request, 'No table structure detected. Using fallback extraction.')
                        template_processor = TemplateProcessor(ocr_engine)
                        extracted_fields = template_processor.process_document_with_template(
                            full_path, document.template.structure or {}
                        )
                        document.extracted_data = {
                            'fields': [
                                {
                                    'name': field.name,
                                    'value': field.value,
                                    'confidence': field.confidence
                                } for field in extracted_fields
                            ]
                        }
                        messages.success(request, f'Document reprocessed with template. Extracted {len(extracted_fields)} fields.')
                else:
                    # Old template format - use template processor
                    template_processor = TemplateProcessor(ocr_engine)
                    extracted_fields = template_processor.process_document_with_template(
                        full_path, document.template.structure or {}
                    )
                    document.extracted_data = {
                        'fields': [
                            {
                                'name': field.name,
                                'value': field.value,
                                'confidence': field.confidence
                            } for field in extracted_fields
                        ]
                    }
                    messages.success(request, f'Document reprocessed with template. Extracted {len(extracted_fields)} fields.')
            else:
                # General OCR reprocessing
                ocr_result = ocr_engine.extract_text(full_path)
                document.extracted_data = {
                    'text': ocr_result.text,
                    'confidence': ocr_result.confidence,
                    'engine': ocr_result.engine
                }
                messages.success(request, f'Document reprocessed. Confidence: {ocr_result.confidence:.1f}%')
            
            document.processing_status = 'completed'
            document.save()
            
            return redirect('documents:document_detail', document_id=document_id)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error reprocessing document: {str(e)}')
            return redirect('documents:document_detail', document_id=document_id)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def document_export(request, document_id):
    """Export document as text file"""
    document = get_object_or_404(Document, id=document_id)
    
    # Prepare export content
    if document.template and document.extracted_data and 'fields' in document.extracted_data:
        # Template-based export with structured data
        content_lines = [f"Document: {document.name}"]
        content_lines.append(f"Template: {document.template.name}")
        content_lines.append(f"Processed: {document.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
        content_lines.append("=" * 50)
        content_lines.append("")
        
        for field in document.extracted_data['fields']:
            field_name = field.get('name', 'Unknown Field')
            field_value = field.get('value', '')
            confidence = field.get('confidence', 0)
            content_lines.append(f"{field_name}: {field_value}")
            if confidence:
                content_lines.append(f"  [Confidence: {confidence:.1f}%]")
            content_lines.append("")
        
        export_content = '\n'.join(content_lines)
        filename = f"{document.name}_structured.txt"
        
    else:
        # General text export
        export_content = document.extracted_data.get('text', '') if document.extracted_data else ''
        filename = f"{document.name}_text.txt"
    
    # Create response
    response = HttpResponse(export_content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def document_reextract_field(request, document_id):
    """Re-extract a specific field from the document using OCR"""
    import json
    
    if request.method == 'POST':
        try:
            from ocr_processing.ocr_core import OCREngine, TemplateProcessor
            from django.conf import settings
            import os
            
            document = get_object_or_404(Document, id=document_id)
            
            # Get field name from request
            data = json.loads(request.body)
            field_name = data.get('field_name')
            
            if not field_name:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Field name is required'
                }, status=400)
            
            # Check if document has a template
            if not document.template:
                return JsonResponse({
                    'status': 'error',
                    'message': 'This document has no associated template'
                }, status=400)
            
            # Get document file path
            full_path = os.path.join(settings.MEDIA_ROOT, document.file.name)
            
            if not os.path.exists(full_path):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Original document file not found'
                }, status=404)
            
            # Re-extract the field
            ocr_engine = OCREngine()
            template_processor = TemplateProcessor(ocr_engine)
            
            # Extract all fields again using the template
            extracted_fields = template_processor.process_document_with_template(
                full_path,
                document.template.structure
            )
            
            # Find the specific field
            field_value = None
            field_confidence = None
            
            for field in extracted_fields:
                if field.name == field_name:
                    field_value = field.value
                    field_confidence = field.confidence
                    break
            
            if field_value is None:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Field "{field_name}" not found in extraction result'
                }, status=404)
            
            # Update the document's extracted data for this field
            if document.extracted_data and 'fields' in document.extracted_data:
                for i, field in enumerate(document.extracted_data['fields']):
                    if field.get('name') == field_name:
                        document.extracted_data['fields'][i]['value'] = field_value
                        document.extracted_data['fields'][i]['confidence'] = field_confidence
                        break
                
                document.save()
            
            return JsonResponse({
                'status': 'success',
                'field_name': field_name,
                'field_value': field_value,
                'confidence': field_confidence,
                'message': f'Field "{field_name}" re-extracted successfully'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error re-extracting field: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)


def document_download_excel(request, document_id):
    """Download the populated Excel file for a document"""
    document = get_object_or_404(Document, id=document_id)
    
    if not document.excel_file:
        messages.error(request, 'No Excel file available for this document')
        return redirect('documents:document_detail', document_id=document_id)
    
    try:
        from django.conf import settings
        import os
        
        excel_path = os.path.join(settings.MEDIA_ROOT, document.excel_file.name)
        
        if not os.path.exists(excel_path):
            messages.error(request, 'Excel file not found')
            return redirect('documents:document_detail', document_id=document_id)
        
        # Read file and return as response
        with open(excel_path, 'rb') as excel_file:
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{document.name}_data.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
    except Exception as e:
        messages.error(request, f'Error downloading Excel file: {str(e)}')
        return redirect('documents:document_detail', document_id=document_id)


def template_export_all_documents(request, template_id):
    """Export all documents for a template into a single Excel file"""
    template = get_object_or_404(Template, id=template_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.excel_manager import ExcelTemplateManager
            from django.conf import settings
            import os
            from datetime import datetime
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{template.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            if not custom_filename.endswith('.xlsx'):
                custom_filename += '.xlsx'
            
            # Get all documents for this template
            documents = Document.objects.filter(template=template, processing_status='completed')
            
            if not documents.exists():
                messages.error(request, 'No processed documents found for this template')
                return redirect('templates:template_detail', template_id=template_id)
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create consolidated Excel
            excel_manager = ExcelTemplateManager()
            excel_path = excel_manager.create_multi_document_export(
                template, 
                documents, 
                output_path,
                custom_filename
            )
            
            # Return file as download
            with open(excel_path, 'rb') as excel_file:
                response = HttpResponse(
                    excel_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting documents: {str(e)}')
            return redirect('templates:template_detail', template_id=template_id)
    
    # GET request - show export form
    documents = Document.objects.filter(template=template, processing_status='completed')
    context = {
        'template': template,
        'document_count': documents.count()
    }
    return render(request, 'documents/template_export_form.html', context)


def document_export_excel(request, document_id):
    """Export a single document's data to Excel with custom filename"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.excel_manager import ExcelTemplateManager
            from django.conf import settings
            from datetime import datetime
            import os
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{document.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            if not custom_filename.endswith('.xlsx'):
                custom_filename += '.xlsx'
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create Excel file
            excel_manager = ExcelTemplateManager()
            
            if document.template:
                # Create Excel from template
                excel_path = excel_manager.create_multi_document_export(
                    document.template,
                    [document],
                    output_path,
                    custom_filename
                )
            else:
                # Create simple Excel from extracted data
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Extracted Data"
                
                # Write data
                if document.extracted_data and 'text' in document.extracted_data:
                    ws.cell(row=1, column=1, value="Extracted Text")
                    ws.cell(row=2, column=1, value=document.extracted_data['text'])
                
                wb.save(output_path)
                excel_path = output_path
            
            # Return file as download
            with open(excel_path, 'rb') as excel_file:
                response = HttpResponse(
                    excel_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting to Excel: {str(e)}')
            return redirect('documents:document_detail', document_id=document_id)
    
    # GET request - show export form
    return render(request, 'documents/document_export_form.html', {'document': document})


def document_export_docx(request, document_id):
    """Export a single document's data to Word (.docx) with custom filename"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.docx_exporter import DocxExporter
            from django.conf import settings
            from datetime import datetime
            import os
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{document.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
            
            if not custom_filename.endswith('.docx'):
                custom_filename += '.docx'
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create Word document
            docx_exporter = DocxExporter()
            
            if document.template:
                # Export structured template data
                docx_path = docx_exporter.export_template_data_to_docx(
                    document,
                    document.template,
                    output_path
                )
            else:
                # Export plain text
                text = document.extracted_data.get('text', 'No text extracted')
                metadata = {
                    'Document': document.name,
                    'Processed': document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Confidence': f"{document.confidence_score:.1f}%" if document.confidence_score else 'N/A'
                }
                docx_path = docx_exporter.export_plain_text_to_docx(
                    text,
                    output_path,
                    title=document.name,
                    metadata=metadata
                )
            
            # Return file as download
            with open(docx_path, 'rb') as docx_file:
                response = HttpResponse(
                    docx_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting to Word: {str(e)}')
            return redirect('documents:document_detail', document_id=document_id)
    
    # GET request - show export form
    return render(request, 'documents/document_export_docx_form.html', {'document': document})


def document_export_csv(request, document_id):
    """Export a single document to CSV format"""
    import csv
    from io import StringIO
    
    document = get_object_or_404(Document, id=document_id)
    
    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)
    
    if document.template:
        # Template-based export
        field_names = document.template.get_field_names()
        
        # Write header
        writer.writerow(field_names)
        
        # Write data
        extracted_data = document.extracted_data
        if 'fields' in extracted_data:
            values = [field.get('value', '') for field in extracted_data['fields']]
            writer.writerow(values)
        elif 'cells' in extracted_data:
            # Table format - extract ALL data rows (skip header row 0)
            cells = extracted_data['cells']
            # Group cells by row
            row_data = {}
            for cell_data in cells:
                row = cell_data.get('row', 0)
                col = cell_data.get('col', 0)
                text = cell_data.get('text', '')
                if row not in row_data:
                    row_data[row] = {}
                row_data[row][col] = text
            
            # Write all data rows (skip header at row 0)
            for row_idx in sorted(row_data.keys()):
                if row_idx > 0:  # Skip header row
                    values = [row_data[row_idx].get(i, '') for i in range(len(field_names))]
                    writer.writerow(values)
    else:
        # General text export - just document info
        writer.writerow(['Document', 'Extracted Text', 'Confidence'])
        text = document.extracted_data.get('text', '') if document.extracted_data else ''
        confidence = document.confidence_score if document.confidence_score else 'N/A'
        writer.writerow([document.name, text, confidence])
    
    # Create response
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    filename = f"{document.name}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def template_export_all_documents_csv(request, template_id):
    """Export all documents for a template to CSV format"""
    import csv
    from io import StringIO
    
    template = get_object_or_404(Template, id=template_id)
    documents = Document.objects.filter(template=template, processing_status='completed')
    
    if not documents.exists():
        messages.error(request, 'No processed documents found for this template')
        return redirect('templates:template_detail', template_id=template_id)
    
    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)
    
    # Get field names from template
    field_names = template.get_field_names()
    
    # Write header
    header = ['Document Name', 'Processed Date'] + field_names + ['Confidence']
    writer.writerow(header)
    
    # Write data rows
    for document in documents:
        extracted_data = document.extracted_data
        
        if 'fields' in extracted_data:
            # Template-based format - single row per document
            row = [
                document.name,
                document.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            for field in extracted_data['fields']:
                row.append(field.get('value', ''))
            
            # Add confidence
            confidence = f"{document.confidence_score:.1f}%" if document.confidence_score else 'N/A'
            row.append(confidence)
            writer.writerow(row)
            
        elif 'cells' in extracted_data:
            # Table format - multiple rows per document (skip header row 0)
            cells = extracted_data['cells']
            # Group cells by row
            cell_rows = {}
            for cell_data in cells:
                cell_row = cell_data.get('row', 0)
                col = cell_data.get('col', 0)
                text = cell_data.get('text', '')
                if cell_row not in cell_rows:
                    cell_rows[cell_row] = {}
                cell_rows[cell_row][col] = text
            
            # Export all data rows (skip header at row 0)
            for data_row_idx in sorted(cell_rows.keys()):
                if data_row_idx > 0:  # Skip header row
                    row = [
                        document.name,
                        document.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    ]
                    for i in range(len(field_names)):
                        row.append(cell_rows[data_row_idx].get(i, ''))
                    
                    # Add confidence
                    confidence = f"{document.confidence_score:.1f}%" if document.confidence_score else 'N/A'
                    row.append(confidence)
                    
                    writer.writerow(row)
        else:
            # No data - write empty row
            row = [
                document.name,
                document.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            row.extend([''] * len(field_names))
            confidence = f"{document.confidence_score:.1f}%" if document.confidence_score else 'N/A'
            row.append(confidence)
            writer.writerow(row)
    
    # Create response
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    filename = f"{template.name}_export.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def template_export_all_documents_docx(request, template_id):
    """Export all documents for a template into a single Word file"""
    template = get_object_or_404(Template, id=template_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.docx_exporter import DocxExporter
            from django.conf import settings
            import os
            from datetime import datetime
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{template.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
            
            if not custom_filename.endswith('.docx'):
                custom_filename += '.docx'
            
            # Get all documents for this template
            documents = Document.objects.filter(template=template, processing_status='completed')
            
            if not documents.exists():
                messages.error(request, 'No processed documents found for this template')
                return redirect('templates:template_detail', template_id=template_id)
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create consolidated Word document
            docx_exporter = DocxExporter()
            docx_path = docx_exporter.export_multiple_documents_to_docx(
                documents,
                template,
                output_path
            )
            
            # Return file as download
            with open(docx_path, 'rb') as docx_file:
                response = HttpResponse(
                    docx_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting documents to Word: {str(e)}')
            return redirect('templates:template_detail', template_id=template_id)
    
    # GET request - show export form
    documents = Document.objects.filter(template=template, processing_status='completed')
    context = {
        'template': template,
        'document_count': documents.count()
    }
    return render(request, 'documents/template_export_docx_form.html', context)


def document_export_pdf(request, document_id):
    """Export a single document's data to PDF with custom filename"""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.pdf_filler import PDFFiller
            from django.conf import settings
            from datetime import datetime
            import os
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{document.name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            if not custom_filename.endswith('.pdf'):
                custom_filename += '.pdf'
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create PDF
            pdf_filler = PDFFiller()
            
            if document.template:
                # Export structured template data as PDF
                pdf_path = pdf_filler.create_pdf_from_template_data(
                    document,
                    document.template,
                    output_path
                )
            else:
                # Export plain text as PDF
                text = document.extracted_data.get('text', 'No text extracted')
                metadata = {
                    'Document': document.name,
                    'Processed': document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Confidence': f"{document.confidence_score:.1f}%" if document.confidence_score else 'N/A'
                }
                pdf_path = pdf_filler.create_pdf_from_text(
                    text,
                    output_path,
                    title=document.name,
                    metadata=metadata
                )
            
            # Return file as download
            with open(pdf_path, 'rb') as pdf_file:
                response = HttpResponse(
                    pdf_file.read(),
                    content_type='application/pdf'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting to PDF: {str(e)}')
            return redirect('documents:document_detail', document_id=document_id)
    
    # GET request - show export form
    return render(request, 'documents/document_export_pdf_form.html', {'document': document})


def template_export_all_documents_pdf(request, template_id):
    """Export all documents for a template into a single PDF file"""
    template = get_object_or_404(Template, id=template_id)
    
    if request.method == 'POST':
        try:
            from ocr_processing.pdf_filler import PDFFiller
            from django.conf import settings
            import os
            from datetime import datetime
            
            # Get custom filename from user
            custom_filename = request.POST.get('export_filename', '')
            if not custom_filename:
                custom_filename = f"{template.name}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            if not custom_filename.endswith('.pdf'):
                custom_filename += '.pdf'
            
            # Get all documents for this template
            documents = Document.objects.filter(template=template, processing_status='completed')
            
            if not documents.exists():
                messages.error(request, 'No processed documents found for this template')
                return redirect('templates:template_detail', template_id=template_id)
            
            # Create output path
            output_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, custom_filename)
            
            # Create consolidated PDF
            pdf_filler = PDFFiller()
            pdf_path = pdf_filler.create_consolidated_pdf(
                documents,
                template,
                output_path
            )
            
            # Return file as download
            with open(pdf_path, 'rb') as pdf_file:
                response = HttpResponse(
                    pdf_file.read(),
                    content_type='application/pdf'
                )
                response['Content-Disposition'] = f'attachment; filename="{custom_filename}"'
                return response
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error exporting documents to PDF: {str(e)}')
            return redirect('templates:template_detail', template_id=template_id)
    
    # GET request - show export form
    documents = Document.objects.filter(template=template, processing_status='completed')
    context = {
        'template': template,
        'document_count': documents.count()
    }
    return render(request, 'documents/template_export_pdf_form.html', context)


def serve_document_file(request, document_id):
    """Serve document file from database"""
    from basemode.file_storage import serve_file_from_db
    
    document = get_object_or_404(Document, id=document_id)
    
    if not document.file_data:
        # Fallback to file system if no DB data
        if document.file:
            from django.http import FileResponse
            return FileResponse(document.file.open(), content_type=document.file_type or 'application/octet-stream')
        return HttpResponse('File not found', status=404)
    
    return serve_file_from_db(
        document.file_data,
        document.file_name or 'document_file',
        document.file_type or 'application/octet-stream',
        as_attachment=False
    )


def serve_document_excel(request, document_id):
    """Serve Excel file from database"""
    from basemode.file_storage import serve_file_from_db
    
    document = get_object_or_404(Document, id=document_id)
    
    if not document.excel_data:
        return HttpResponse('Excel file not found', status=404)
    
    return serve_file_from_db(
        document.excel_data,
        document.excel_name or f'{document.name}_extracted.xlsx',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True
    )
